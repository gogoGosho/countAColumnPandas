import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("combined.xlsx")
result = 0

for i in range(len(wb.sheetnames)):
    df = pd.read_excel("combined.xlsx", i)
    chars = df["Characters"].sum()
    result += chars
print(result)

# res = len(wb.sheetnames)

# print(df['Characters'].sum())